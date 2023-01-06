import openpyxl
from openpyxl.utils import get_column_letter

file = 'questions_new - Copy.xlsx'
quiz_dict = {}
wbook = openpyxl.load_workbook(file)
sheet_q = wbook['questions']
num_rows_q = sheet_q.max_row
count_q = 0
for row in range(2, (num_rows_q + 1)):
    num_ans = sheet_q['D' + str(row)].value
    theme = sheet_q['A' + str(row)].value
    answers = []
    for num in range(5, (num_ans + 5)):
        ans = str(sheet_q[get_column_letter(num) + str(row)].value)
        if len(ans) > 100:
            print(row, len(ans), get_column_letter(num))
