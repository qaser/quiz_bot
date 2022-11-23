import os
import openpyxl

from openpyxl.utils import get_column_letter
from config.mongo_config import questions
from config.bot_config import bot, dp
from config.telegram_config import ADMIN_TELEGRAM_ID
from utils.constants import THEMES
from utils.utils import word_conjugate
from aiogram import Dispatcher, types


async def import_file(message: types.Message):
    await message.answer('Отправьте следующим сообщением файл с вопросами')


@dp.message_handler(content_types=types.ContentType.DOCUMENT)
async def download_file(message):
    file_name = message.document.file_name
    if file_name.endswith('.xlsx'):
        file_id = message.document.file_id
        file = await bot.get_file(file_id)
        file_path = file.file_path
        await bot.download_file(
            file_path=file_path,
            destination=r'static/questions.xlsx'
        )
        await import_from_excel(message.from_user.id)
    else:
        await bot.answer(
            'Данный тип файлов не поддерживается.\n Скачайте шаблон /example'
        )


async def import_from_excel(user_id):
    wbook = openpyxl.load_workbook('static/questions.xlsx')
    sheet = wbook['questions']
    num_rows = sheet.max_row
    count_q = 0
    for row in range(2, (num_rows + 1)):
        try:
            num_ans = sheet['D' + str(row)].value
            theme = sheet['A' + str(row)].value
            answers = []
            for num in range(5, (num_ans + 5)):
                ans = sheet[get_column_letter(num) + str(row)].value
                if ans == '' or ans == None or ans == 'null' or len(ans) > 200:
                    break
                answers.append(ans)
            if len(answers) < num_ans:
                continue
            questions.insert_one({
                'theme': theme,
                'question': sheet['B' + str(row)].value,
                'correct_answer': sheet['C' + str(row)].value,
                'num_answers': num_ans,
                'answers': answers,
            })
            count_q += 1
            new_themes = []
            if theme not in THEMES.keys():
                new_themes.append(theme)
        except:
            continue
    if len(new_themes) > 0:
        await bot.send_message(
            ADMIN_TELEGRAM_ID,
            f'Добавлены вопросы с новыми темами: {new_themes}'
        )
    os.remove('static/questions.xlsx')
    words_q = ['вопрос', 'вопроса', 'вопросов']
    words_z = ['Загружен', 'Загружено', 'Загружено']
    text_q = word_conjugate(count_q, words_q)
    text_z = word_conjugate(count_q, words_z)
    await bot.send_message(
        chat_id=user_id,
        text=f'Файл получен.\n{text_z} {count_q} {text_q} из {num_rows-1}.'
    )


async def send_example(message: types.Message):
    await message.answer_document(open('static/example/questions.xlsx', 'rb'))



def register_handlers_excel(dp: Dispatcher):
    dp.register_message_handler(import_file, commands='import')
    dp.register_message_handler(send_example, commands='example')
